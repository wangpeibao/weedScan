# 获取定位页面
# 结合套用页面
import os
from datetime import datetime
import openpyxl

from PyQt5.QtCore import QUrl, QTimer
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QLineEdit, \
    QGridLayout, QMessageBox
from PyQt5.QtWebEngineWidgets import *

from widget.UploadCard import UploadCard


class Location(QWidget):
    def __init__(self, base_path, username):
        super(Location, self).__init__()
        self.base_path = base_path
        self.username = username
        # 定位获取的信息
        self.location = {}
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.get_location)

        self.browser = QWebEngineView()
        self.browser.settings().setAttribute(QWebEngineSettings.AllowRunningInsecureContent, True)
        if "home" in self.base_path:
            url = "file:///home/wang/html_tmp/location.html"
        else:
            url = "file:///C:/html_tmp/location.html"

        print('---------', url)
        self.browser.load(QUrl(url))

        self.edit_search = QLineEdit()
        self.edit_search.setPlaceholderText("查询地点")
        self.btn_search = QPushButton("查询")
        self.btn_search.clicked.connect(self.search_location)
        self.edit_address = QLineEdit()
        self.btn_submit = QPushButton("确定")
        self.filename = None

        layout = QGridLayout()
        layout.addWidget(self.browser, 0, 0, 1, 2)
        layout.addWidget(self.edit_search, 1, 0)
        layout.addWidget(self.btn_search, 1, 1)
        layout.addWidget(self.edit_address, 2, 0, 1, 2)
        layout.addWidget(self.btn_submit, 3, 0, 1, 2)
        self.setLayout(layout)

        # 设置格式
        style_str = "QLabel{font-size: 25px;}" + "QLineEdit{font-size: 25px;}" + \
                    "QPushButton{font-size: 25px; background-color: green; min-height: 50px}"
        self.setStyleSheet(style_str)
        self.timer.start(1000)

    def search_location(self):
        keyword = self.edit_search.text()
        if keyword:
            self.browser.page().runJavaScript('search_place("%s");' % keyword)

    def get_location(self):
        self.browser.page().runJavaScript('get_location();', self.get_callback)

    def get_callback(self, result):
        self.location = result
        if result:
            self.edit_address.setText(result["address"])

    def js_callback(self, result):
        # 进行文件的创建
        wb = openpyxl.Workbook()
        wb.create_sheet("基础信息", 0)
        wb.create_sheet("现场照片", 1)
        wb.create_sheet("许可证照片", 2)
        wb.create_sheet("违规烟草记录", 3)
        ws = wb["基础信息"]
        ws["A1"] = "时间"
        ws["B1"] = "地点"
        ws["C1"] = "操作人"
        ws["D1"] = "定位地点"
        ws["E1"] = "经度"
        ws["F1"] = "纬度"
        ws["A2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["B2"] = self.edit_address
        ws["C2"] = self.username
        if result:
            ws["D2"] = result["address"]
            ws["E2"] = result["lng"]
            ws["F2"] = result["lat"]
        else:
            ws["D2"] = "未知地点"
            ws["E2"] = "0.0"
            ws["F2"] = "0.0"
        wb.save(self.filename)

    # 获取地点信息和经纬度
    def handle_info(self):
        self.timer.stop()
        address = self.edit_address.text()
        if not address:
            QMessageBox.critical(self, "错误", "请填写地址信息", QMessageBox.Yes)
            return None
        # 处理地理信息，并且创建响应的文件夹
        now_time_str1 = datetime.now().strftime("%H%M")
        self.base_path += address + now_time_str1 + "/"
        try:
            os.makedirs(self.base_path)
        except Exception as e:
            print(e)
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(title="基础信息")
        ws["A1"] = "时间"
        ws["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["A2"] = "地点"
        ws["B2"] = address
        ws["A3"] = "登录人"
        ws["B3"] = self.username
        ws["A4"] = "定位地点"
        if self.location:
            ws["B4"] = self.location["address"]
            ws["A5"] = "经度"
            ws["B5"] = self.location["lng"]
            ws["A6"] = "纬度"
            ws["B6"] = self.location["lat"]
        else:
            ws["B4"] = "未知地点"
            ws["A5"] = "经度"
            ws["B5"] = "0.0"
            ws["A6"] = "纬度"
            ws["B6"] = "0.0"
        wb.save(self.base_path + "案件信息.xlsx")
        return self.base_path

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ui = Location(123)
    ui.show()
    sys.exit(app.exec_())