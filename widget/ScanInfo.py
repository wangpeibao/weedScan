# 扫描信息的展示窗口
import re
import sqlite3
from enum import IntEnum

import openpyxl
from PyQt5.QtGui import QIntValidator
from PyQt5.QtWidgets import QTableWidget, QWidget, QHBoxLayout, QApplication, QHeaderView, QPushButton, QVBoxLayout, \
    QComboBox, QLineEdit, QTableWidgetItem, QLabel, QMessageBox

from widget.Function import getLetter, get_yan_info
from widget.ScanDial import ScanDial
from widget.ScanHand import ScanHand
from widget.ScanSwitch import ScanSwitch


class ColIndex(IntEnum):
    yan_id = 0
    yan_name = 1
    yan_pinzhong = 2
    yan_price = 3
    yan_unit = 4
    yan_sort = 5
    yan_baozhuang = 6
    yan_peisong_status = 7
    yan_peisong_code = 8
    yan_count = 9
    yan_total = 10
    yan_del = 11

# 继承按钮控件
class TableBtn(QPushButton):
    def __init__(self, name, index):
        super(TableBtn, self).__init__(name)
        self.index = index

    def set_index(self, index):
        self.index = index

# 集成输入框控件
class TableEdit(QLineEdit):
    def __init__(self, index):
        super(TableEdit, self).__init__()
        self.index = index

    def set_index(self, index):
        self.index = index


class ScanWidget(QWidget):
    def __init__(self, filename):
        super(ScanWidget, self).__init__()
        # 窗口类大小
        self.filename = filename

        # 打开数据库
        database = sqlite3.connect("./app.db")
        self.query = database.cursor()

        # 扫描录入和手动录入btn
        scanInputBtn = QPushButton("扫描录入")
        scanInputBtn.clicked.connect(self.showScanDialog)
        handInputBtn = QPushButton("手动录入")
        handInputBtn.clicked.connect(self.showScanHand)

        # 总价label和完成btn
        self.totalLabel = QLabel("总价: %.2f" % 0)
        self.finishBtn = QPushButton("完成")
        # self.finishBtn.clicked.connect(self.uploadData)

        # 表格空间
        self.table_info = QTableWidget()
        self.table_info.setColumnCount(12)
        self.table_info.setHorizontalHeaderLabels(["条码", "名称", "品种", "单价", "单位", "分类", "包装完整性", "有无配送码", "配送码", "数量", "总价", "操作"])
        self.table_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 布局填充数据
        layout = QVBoxLayout()
        btn_layout = QHBoxLayout()
        btn_widget = QWidget()
        btn_layout.addWidget(scanInputBtn)
        btn_layout.addWidget(handInputBtn)
        btn_widget.setLayout(btn_layout)
        layout.addWidget(self.table_info)
        layout.addWidget(btn_widget)
        layout.addWidget(self.totalLabel)
        layout.addWidget(self.finishBtn)
        self.setLayout(layout)

        # 绘制页面格式
        style_str = "QLabel{font-size: 30px;}" + "QLineEdit{font-size: 30px;}" + \
                    "QPushButton{font-size: 25px; background-color: green; min-height: 30px}" + \
                    "QComboBox{font-size: 30px;}" + \
                    "QHeaderView{font-size: 25px;} QTableWidget{font-size: 25px;}"
        self.setStyleSheet(style_str)

        self.initYanData()

    # 初始化数据
    def initYanData(self):
        wb = openpyxl.load_workbook(self.filename)
        results = get_yan_info(wb)
        for result in results:
            self.tableAddInfo(result)



    # 显示弹窗扫描录入
    def showScanDialog(self):
        self.scan_dial = ScanDial()
        self.scan_dial.dialogSignal.connect(self.queryYanInfo)
        self.scan_dial.show()

    # 显示弹窗选择录入
    def showScanSwitch(self, yan_id):
        self.scan_switch = ScanSwitch(yan_id)
        self.scan_switch.dialogSignal.connect(self.handleSwitch)
        self.scan_switch.show()

    # 显示弹窗手动录入
    def showScanHand(self, yan_id):
        self.scan_hand = ScanHand(yan_id)
        # self.scan_hand.yanidSignal.connect(self.queryYanInfo)
        self.scan_hand.yaninfoSignal.connect(self.handleHand)
        self.scan_hand.show()

    # 查询烟草信息
    def queryYanInfo(self, yan_id):
        res = self.querySql(yan_id)  # 如果未查询到数据为空字典
        self.scan_dial.close()
        if not res:  # 未查询到数据，弹窗让用户选择重新扫描还是手动输入
            self.showScanSwitch(yan_id)
        else:  # 查询到数据，显示到页面中
            self.tableAddInfo(res)

    # 封装查询语句
    def querySql(self, yan_id):
        result = self.query.execute("select * from yaninfo where yan_id == %s" % yan_id)
        response = {}
        for row in result:
            response = {
                "yan_id": row[0],
                "yan_name": row[1],
                "yan_type": row[2],
                "yan_price": str(row[3]),
                "yan_unit": "条",
                "yan_pinzhong": "卷烟"
            }
        return response

    # 选择窗口
    def handleSwitch(self, data, yan_id):
        self.scan_switch.close()
        if data == 0:  # 继续扫描
            self.showScanDialog()
        else:  # 手动输入
            print(123, yan_id)
            self.showScanHand(yan_id)

    # 处理手写数据
    def handleHand(self, data):
        self.scan_hand.db.close()
        self.scan_hand.close()
        self.tableAddInfo(data)

    # 设置窗口表格数据
    def tableAddInfo(self, data):
        # 如果已经在数据中,提示
        if self.idHasIn(data["yan_id"]):
            QMessageBox.critical(self, "错误", "数据已经在列表中", QMessageBox.Yes)
            return

        index = self.table_info.rowCount()
        self.table_info.setRowCount(index + 1)
        self.table_info.setItem(index, ColIndex.yan_id, QTableWidgetItem(data["yan_id"]))
        self.table_info.setItem(index, ColIndex.yan_name, QTableWidgetItem(data["yan_name"]))
        yan_pinzhong = QLineEdit()
        if "yan_pinzhong" not in data.keys():
            data["yan_pinzhong"] = "卷烟"
        yan_pinzhong.setText(data["yan_pinzhong"])
        self.table_info.setCellWidget(index, ColIndex.yan_pinzhong, yan_pinzhong)
        self.table_info.setItem(index, ColIndex.yan_price, QTableWidgetItem(str(data["yan_price"])))
        self.table_info.setItem(index, ColIndex.yan_unit, QTableWidgetItem(data["yan_unit"]))
        # 包装完整性
        yan_baozhuang = QComboBox()
        yan_baozhuang.addItems(["完好无破损", "有破损"])
        yan_baozhuang.setCurrentIndex(0)
        if "yan_baozhuang" in data.keys():
            if data["yan_baozhuang"] == "完好无破损":
                yan_baozhuang.setCurrentIndex(0)
            else:
                yan_baozhuang.setCurrentIndex(1)
        else:
            yan_baozhuang.setCurrentIndex(0)
        self.table_info.setCellWidget(index, ColIndex.yan_baozhuang, yan_baozhuang)
        # 烟草分类
        yan_sort = QComboBox()
        yan_sort.addItems(["假", "非", "无专卖字样", "授权生产", "专供出口"])
        # 如果yan_id不是以6901028开头的，默认是走私烟
        if re.match(r"^6901028.*", data["yan_id"]):
            pass
        else:
            data["yan_sort"] = "无专卖字样"
        if "yan_sort" in data.keys():
            if data["yan_sort"] == "假":
                yan_sort.setCurrentIndex(0)
            elif data["yan_sort"] == "非":
                yan_sort.setCurrentIndex(1)
            elif data["yan_sort"] == "无专卖字样":
                yan_sort.setCurrentIndex(2)
            elif data["yan_sort"] == "授权生产":
                yan_sort.setCurrentIndex(3)
            else:
                yan_sort.setCurrentIndex(4)
        self.table_info.setCellWidget(index, ColIndex.yan_sort, yan_sort)
        # 有无配送码
        peisong_code = QLineEdit()
        peisong_status = QComboBox()
        peisong_status.addItems(["有配送码", "无配送码", "配送码模糊不清"])
        if "yan_peisong_status" in data.keys():
            if data["yan_peisong_status"] == "配送码":
                peisong_status.setCurrentIndex(0)
                peisong_code.setText(data["peisong_code"])
            elif data["yan_peisong_status"] == "无配送码":
                peisong_status.setCurrentIndex(1)
            else:
                peisong_status.setCurrentIndex(2)
        else:
            peisong_status.setCurrentIndex(0)
        self.table_info.setCellWidget(index, ColIndex.yan_peisong_status, peisong_status)
        self.table_info.setCellWidget(index, ColIndex.yan_peisong_code, peisong_code)
        # 数量
        yan_count = TableEdit(index=index)
        if "yan_count" in data.keys():
            yan_count.setText(data["yan_count"])
        yan_validator = QIntValidator(1, 9999999)
        yan_count.setValidator(yan_validator)
        yan_count.textChanged.connect(self.getSelfTotalPrice)
        self.table_info.setCellWidget(index, ColIndex.yan_count, yan_count)
        # 单种类烟草总价
        if "yan_total" in data.keys():
            self.table_info.setItem(index, ColIndex.yan_total, QTableWidgetItem(data["yan_total"]))
        else:
            self.table_info.setItem(index, ColIndex.yan_total, QTableWidgetItem(0))
        # 删除按钮
        btn_del = TableBtn(name="删除", index=index)
        btn_del.clicked.connect(self.deleteRow)
        self.table_info.setCellWidget(index, ColIndex.yan_del, btn_del)

    # 获取单条烟草的总价
    def getSelfTotalPrice(self):
        try:
            count = int(self.sender().text())
        except Exception as e:
            count = 0
        index = self.sender().index
        price = float(self.table_info.item(index, ColIndex.yan_price.value).text())
        self.table_info.item(self.sender().index, ColIndex.yan_total).setText("%.2f" % (count * price))
        self.totalLabel.setText("总价: %.2f" % self.getAllTotalPrice())

    # 获取此次扫描的总价
    def getAllTotalPrice(self):
        row_count = self.table_info.rowCount()
        price_count = 0
        for i in range(row_count):
            try:
                self_count = float(self.table_info.item(i, ColIndex.yan_total).text())
            except Exception as e:
                print(e)
                self_count = 0
            price_count += self_count
        return price_count


    # 删除所在行
    def deleteRow(self):
        sender = self.sender()
        self.table_info.removeRow(sender.index)
        self.updateRowIndex()

    # 更新索引
    def updateRowIndex(self):
        row_count = self.table_info.rowCount()
        for i in range(row_count):
            self.table_info.cellWidget(i, ColIndex.yan_count.value).set_index(i)
            self.table_info.cellWidget(i, ColIndex.yan_del.value).set_index(i)

    # 判断列表中是否有该数据
    def idHasIn(self, id):
        row_count = self.table_info.rowCount()
        has_in = False
        for i in range(row_count):
            if self.table_info.item(i, ColIndex.yan_id.value).text() == id:
                has_in = True
                break
        return has_in

    # 上传数据
    def uploadData(self):
        wb = openpyxl.load_workbook(self.filename)
        try:
            ws = wb["违规烟草记录"]
        except Exception as e:
            ws = wb.create_sheet(title="违规烟草记录")
        title_list = ["条形码", "名称", "单价", "单位", "分类", "数量", "总价", "品种", "包装完整性", "有无配送码", "配送码"]
        # 设置表头
        for index, title in enumerate(title_list):
            ws[getLetter(index) + "1"] = title
        # 填充数据

        row_count = self.table_info.rowCount()
        num_count = 0
        price_count = 0.0
        for i in range(row_count):
            ws["A" + str(i + 2)] = self.table_info.item(i, ColIndex.yan_id.value).text()
            ws["B" + str(i + 2)] = self.table_info.item(i, ColIndex.yan_name.value).text()
            ws["C" + str(i + 2)] = self.table_info.item(i, ColIndex.yan_price.value).text()
            ws["D" + str(i + 2)] = self.table_info.item(i, ColIndex.yan_unit.value).text()
            yan_sort = self.table_info.cellWidget(i, ColIndex.yan_sort.value).currentText()
            yan_count = self.table_info.cellWidget(i, ColIndex.yan_count.value).text()
            yan_total = self.table_info.item(i, ColIndex.yan_total.value).text()
            yan_pinzhong = self.table_info.cellWidget(i, ColIndex.yan_pinzhong.value).text()

            ws["E" + str(i + 2)] = yan_sort
            ws["F" + str(i + 2)] = yan_count
            num_count += int(yan_count)
            ws["G" + str(i + 2)] = yan_total
            ws["H" + str(i + 2)] = yan_pinzhong
            baozhuang = self.table_info.cellWidget(i, ColIndex.yan_baozhuang.value).currentText()
            ws["I" + str(i + 2)] = baozhuang
            peisong_status = self.table_info.cellWidget(i, ColIndex.yan_peisong_status.value).currentText()
            ws["J" + str(i + 2)] = peisong_status
            peisong_code = self.table_info.cellWidget(i, ColIndex.yan_peisong_code.value).text()
            ws["K" + str(i + 2)] = peisong_code
            price_count += float(yan_total)
        ws["L1"] = "总条数:" + str(num_count)
        ws["M1"] = "总案值:" + str(price_count)
        wb.save(self.filename)



if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ui = ScanWidget()
    ui.show()
    sys.exit(app.exec_())