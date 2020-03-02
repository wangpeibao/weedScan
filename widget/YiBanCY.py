from datetime import datetime

import openpyxl
from PyQt5.QtWidgets import QWidget, QGridLayout, QLabel, QDateTimeEdit, QLineEdit, QComboBox, QTableWidget, \
    QHeaderView, QPushButton, QDateEdit, QTableWidgetItem, QMessageBox, QVBoxLayout, QCheckBox, QTextEdit

from widget.AnYou import AnYou
from widget.Function import get_yan_info, get_chou_data, createYiBan, get_info_from_idcard


class YiBanCY(QWidget):
    def __init__(self, base_path):
        super(YiBanCY, self).__init__()
        self.filename = base_path + "案件信息.xlsx"
        self.base_path = base_path
        style_str = "QLabel{font-size: 30px;}" + "QLineEdit{font-size: 30px;}" + \
                    "QPushButton{font-size: 25px; background-color: green; min-height: 35px}" + \
                    "QComboBox{font-size: 30px;}" + "QCheckBox{font-size: 30px;}" + \
                    "QHeaderView{font-size: 25px;} QTableWidget{font-size: 25px;}" + \
                    "QDateTimeEdit{font-size: 30px;} QMessageBox{font-size: 30px;} QTextEdit{font-color: red}"
        self.setStyleSheet(style_str)
        init_data = self.get_init_info()
        layout = QGridLayout()
        # 开始时间
        self.label_start_time = QLabel("开始时间")
        if init_data["start_time"]:
            self.edit_start_time = QDateTimeEdit(datetime.strptime(init_data["start_time"], "%Y-%m-%d %H:%M"))
        else:
            self.edit_start_time = QDateTimeEdit(datetime.now())
        self.edit_start_time.setDisplayFormat("yyyy-MM-dd HH:mm")
        layout.addWidget(self.label_start_time, 0, 0)
        layout.addWidget(self.edit_start_time, 0, 1)
        # 结束时间
        self.label_end_time = QLabel("结束时间")
        if init_data["end_time"]:
            self.edit_end_time = QDateTimeEdit(datetime.strptime(init_data["end_time"], "%Y-%m-%d %H:%M"))
        else:
            self.edit_end_time = QDateTimeEdit(datetime.now())
        self.edit_end_time.setDisplayFormat("yyyy-MM-dd HH:mm")
        layout.addWidget(self.label_end_time, 0, 2)
        layout.addWidget(self.edit_end_time, 0, 3)
        # 被检查人(工商户字号，公司名称)
        self.label_beijiancharen = QLabel("被检查人(工商户字号，公司名称)")
        if init_data["beijiancharen"]:
            self.edit_beijiancharen = QLineEdit(init_data["beijiancharen"])
        else:
            self.edit_beijiancharen = QLineEdit()
        layout.addWidget(self.label_beijiancharen, 1, 0)
        layout.addWidget(self.edit_beijiancharen, 1, 1)
        # 联系方式
        self.label_lianxifangshi = QLabel("联系方式")
        if init_data["lianxifangshi"]:
            self.edit_lianxifangshi = QLineEdit(init_data["lianxifangshi"])
        else:
            self.edit_lianxifangshi = QLineEdit()
        layout.addWidget(self.label_lianxifangshi, 1, 2)
        layout.addWidget(self.edit_lianxifangshi, 1, 3)
        # 经营者姓名(法定代表人)
        self.label_jingyingzhe = QLabel("经营者姓名(法定代表人)")
        if init_data["jingyingzhe"]:
            self.edit_jingyingzhe = QLineEdit(init_data["jingyingzhe"])
        else:
            self.edit_jingyingzhe = QLineEdit()
        layout.addWidget(self.label_jingyingzhe, 2, 0)
        layout.addWidget(self.edit_jingyingzhe, 2, 1)
        # 身份证件号码
        self.label_shenfenzheng = QLabel("身份证件号码")
        if init_data["shenfenzheng"]:
            self.edit_shenfenzheng = QLineEdit(init_data["shenfenzheng"])
        else:
            self.edit_shenfenzheng = QLineEdit()
        layout.addWidget(self.label_shenfenzheng, 2, 2)
        layout.addWidget(self.edit_shenfenzheng, 2, 3)
        # 身份证件地址
        self.label_shenfenzheng_dizhi = QLabel("身份证件地址")
        if init_data["shenfenzheng_dizhi"]:
            self.edit_shenfenzheng_dizhi = QLineEdit(init_data["shenfenzheng_dizhi"])
        else:
            self.edit_shenfenzheng_dizhi = QLineEdit()
        layout.addWidget(self.label_shenfenzheng_dizhi, 3, 0)
        layout.addWidget(self.edit_shenfenzheng_dizhi, 3, 1)
        # 烟草专卖许可证号码
        self.label_xukezheng = QLabel("烟草专卖许可证号码")
        if init_data["xukezheng"]:
            self.edit_xukezheng = QLineEdit(init_data["xukezheng"])
        else:
            self.edit_xukezheng = QLineEdit()
        layout.addWidget(self.label_xukezheng, 3, 2)
        layout.addWidget(self.edit_xukezheng, 3, 3)
        # 零售户类型
        self.label_lingshouhu = QLabel("零售户类型")
        self.edit_lingshouhu = QComboBox()
        self.edit_lingshouhu.addItems(["个体零售户", "企业零售户"])
        if init_data["lingshouhu"]:
            self.edit_lingshouhu.setCurrentText(init_data["lingshouhu"])
        layout.addWidget(self.label_lingshouhu, 4, 0)
        layout.addWidget(self.edit_lingshouhu, 4, 1)
        # 举报时间(非举报不填写)
        self.jvbao_time = QLabel("举报时间(非举报不填写)")
        if init_data["jvbao_time"]:
            self.edit_jvbao_time = QDateTimeEdit(datetime.strptime(init_data["jvbao_time"], "%Y-%m-%d %H:%M"))
        else:
            self.edit_jvbao_time = QDateTimeEdit(datetime.strptime("2000-01-01 00:00", "%Y-%m-%d %H:%M"))
        self.edit_jvbao_time.setDisplayFormat("yyyy-MM-dd HH:mm")
        layout.addWidget(self.jvbao_time, 4, 2)
        layout.addWidget(self.edit_jvbao_time, 4, 3)
        # 案发时间
        self.label_qita_guanxiren = QLabel("其他关系人")
        if init_data["qita_guanxiren"]:
            self.edit_qita_guanxiren = QLineEdit(init_data["qita_guanxiren"])
        else:
            self.edit_qita_guanxiren = QLineEdit("店员")
        layout.addWidget(self.label_qita_guanxiren, 5, 0)
        layout.addWidget(self.edit_qita_guanxiren, 5, 1)
        # 执法人员
        self.label_zhifarenyuan = QLabel("执法人员")
        if init_data["zhifarenyuan"]:
            self.edit_zhifarenyuan = QLineEdit(init_data["zhifarenyuan"])
        else:
            self.edit_zhifarenyuan = QLineEdit()
        layout.addWidget(self.label_zhifarenyuan, 5, 2)
        layout.addWidget(self.edit_zhifarenyuan, 5, 3)
        # 许可证有效期
        self.label_xukezheng_end = QLabel("许可证有效期")
        if init_data["xukezheng_end"]:
            self.edit_xukezheng_end = QDateEdit(datetime.strptime(init_data["xukezheng_end"], "%Y-%m-%d").date())
        else:
            self.edit_xukezheng_end = QDateEdit(datetime.now().date())
        self.edit_xukezheng_end.setDisplayFormat("yyyy-MM-dd")
        layout.addWidget(self.label_xukezheng_end, 6, 0)
        layout.addWidget(self.edit_xukezheng_end, 6, 1)
        # 许可证地址
        self.label_location = QLabel("许可证地址")
        if init_data["location"]:
            self.edit_location = QLineEdit(init_data["location"])
        else:
            self.edit_location = QLineEdit()
        layout.addWidget(self.label_location, 6, 2)
        layout.addWidget(self.edit_location, 6, 3)
        # 案由选择
        self.label_anyou = QLabel("案由")
        self.edit_anyou = AnYou(init_data["anyou"])
        layout.addWidget(self.label_anyou, 7, 0)
        layout.addWidget(self.edit_anyou, 7, 1, 1, 3)
        # 抽样规则
        rule_str = '<p><font color="red">1.数量不足两条：抽取一条或全部数量形成试样。</font></p>'
        rule_str += '<p><font color="red">2.数量2件以下：所有样品中抽取1~2条形成试样。</font></p>'
        rule_str += '<p><font color="red">3.数量5件以下：每件抽取1~2条，形成样本，再从样本中随机抽取2条形成试样。</font></p>'
        rule_str += '<p><font color="red">4.数量5件~10件：每件中随机抽取1条，形成样本，在从样本中随机抽取2条形成试样。</font></p>'
        rule_str += '<p><font color="red">5.数量10件~50件：随机抽取10件，从每件中随机抽取1条，形成样本，再从样本中随机抽取2~5条形成试样。</font></p>'
        rule_str += '<p><font color="red">6.数量50件以上：随机抽取20件，从每件中随机抽取1条，形成样本，再从样本中随机抽取5~10条形成试样。</font></p>'
        self.chou_rule = QTextEdit()
        self.chou_rule.setHtml(rule_str)
        self.chou_rule.setFixedHeight(130)
        self.label_chou_rule = QLabel("抽样规则")
        layout.addWidget(self.label_chou_rule, 8, 0)
        layout.addWidget(self.chou_rule, 8, 1, 1, 3)
        # 抽样记录表
        self.table_info = QTableWidget()
        self.table_info.setColumnCount(4)
        self.table_info.setHorizontalHeaderLabels(["物品名称", "品牌规格", "查获数量(条)", "抽样数量(条)"])
        self.table_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        if not init_data["chou_data"]:
            wb = openpyxl.load_workbook(self.filename)
            yan_datas = get_yan_info(wb)
            for yan_data in yan_datas:
                init_data["chou_data"].append({
                    "yan_name": yan_data["yan_name"],
                    "yan_count": yan_data["yan_count"],
                    "yan_chou": 1,
                    "yan_pinzhong": yan_data["yan_pinzhong"]
                })
        self.init_table(init_data["chou_data"])
        # 提交按钮
        self.btn_finish = QPushButton("完成")
        layout.addWidget(self.table_info, 9, 0, 1, 4)
        layout.addWidget(self.btn_finish, 10, 0, 1, 4)
        self.setLayout(layout)

    # 获取初始化信息
    def get_init_info(self):
        wb = openpyxl.load_workbook(self.filename)
        try:
            ws = wb["一般案件"]
        except Exception as e:
            print(e)
            print("创建sheet")
            ws = wb.create_sheet(title="一般案件")
        anyou = ['', '', '', '', '']
        if ws["B9"].value:
            anyou[0] = ws["B9"].value
        if ws["C9"].value:
            anyou[1] = ws["C9"].value
        if ws["D9"].value:
            anyou[2] = ws["D9"].value
        if ws["E9"].value:
            anyou[3] = ws["E9"].value
        if ws["E9"].value:
            anyou[4] = ws["F9"].value

        data = {
            "start_time": ws["B1"].value,
            "end_time": ws["B2"].value,
            "beijiancharen": ws["B3"].value,
            "lianxifangshi": ws["B4"].value,
            "jingyingzhe": ws["B5"].value,
            "shenfenzheng": ws["B6"].value,
            "shenfenzheng_dizhi": ws["B7"].value,
            "xukezheng": ws["B8"].value,
            "anyou": anyou,  # B9
            "lingshouhu": ws["B10"].value,
            "jvbao_time": ws["B11"].value,
            "qita_guanxiren": ws["B12"].value,
            "zhifarenyuan": ws["B13"].value,
            "xukezheng_end": ws["B14"].value,
            "location": ws["B15"].value,
            "chou_data": get_chou_data(ws)
        }
        wb.save(self.filename)
        return data

    # 渲染图表
    def init_table(self, datas):
        self.table_info.setRowCount(len(datas))
        for index, data in enumerate(datas):
            self.table_info.setItem(index, 0, QTableWidgetItem(data["yan_pinzhong"]))
            self.table_info.setItem(index, 1, QTableWidgetItem(data["yan_name"]))
            self.table_info.setItem(index, 2, QTableWidgetItem(str(data["yan_count"])))
            self.table_info.setCellWidget(index, 3, QLineEdit(str(data["yan_chou"])))

    # 清空抽样数据
    def clear_chou_data(self, ws):
        index = 16
        while(True):
            yan_name = ws["A" + str(index)].value
            if not yan_name:
                break
            ws["A" + str(index)] = ""
            ws["B" + str(index)] = ""
            ws["C" + str(index)] = ""
            ws["D" + str(index)] = ""

    # 存储抽样数据
    def handle_chou_data(self, ws):
        start_index = 16
        self.clear_chou_data(ws)
        for index in range(self.table_info.rowCount()):
            ws["A" + str(index + start_index)] = self.table_info.item(index, 0).text()
            ws["B" + str(index + start_index)] = self.table_info.item(index, 1).text()
            ws["C" + str(index + start_index)] = self.table_info.item(index, 2).text()
            ws["D" + str(index + start_index)] = self.table_info.cellWidget(index, 3).text()

    # 处理数据
    def handle_info(self):
        result = {}
        start_time = self.edit_start_time.text()
        if not start_time:
            QMessageBox.critical(self, "错误", "请填写开始时间", QMessageBox.Yes)
            return False
        result["start_time"] = datetime.strptime(start_time, "%Y-%m-%d %H:%M")
        end_time = self.edit_end_time.text()
        if not end_time:
            QMessageBox.critical(self, "错误", "请填写结束时间", QMessageBox.Yes)
            return False
        result["end_time"] = datetime.strptime(end_time, "%Y-%m-%d %H:%M")
        beijiancharen = self.edit_beijiancharen.text()
        if not beijiancharen:
            QMessageBox.critical(self, "错误", "请填写被检查人(工商户字号，公司名称)", QMessageBox.Yes)
            return False
        result["beijiancharen"] = beijiancharen
        lianxifangshi = self.edit_lianxifangshi.text()
        if not lianxifangshi:
            QMessageBox.critical(self, "错误", "请填写联系方式", QMessageBox.Yes)
            return False
        result["lianxifangshi"] = lianxifangshi
        jingyingzhe = self.edit_jingyingzhe.text()
        if not jingyingzhe:
            QMessageBox.critical(self, "错误", "请填写经营者姓名(法定代表人)", QMessageBox.Yes)
            return False
        result["jingyingzhe"] = jingyingzhe
        shenfenzheng = self.edit_shenfenzheng.text()
        if not shenfenzheng:
            QMessageBox.critical(self, "错误", "请填写经营者身份证件号", QMessageBox.Yes)
            return False
        else:
            try:
                get_info_from_idcard(shenfenzheng)
            except Exception as e:
                QMessageBox.critical(self, "错误", "请填写正确的经营者身份证号", QMessageBox.Yes)
                return False
        result["shenfenzheng"] = shenfenzheng
        shenfenzheng_dizhi = self.edit_shenfenzheng_dizhi.text()
        if not shenfenzheng_dizhi:
            QMessageBox.critical(self, "错误", "请填写经营者身份证件地址", QMessageBox.Yes)
            return False
        result["shenfenzheng_dizhi"] = shenfenzheng_dizhi
        xukezheng = self.edit_xukezheng.text()
        # if not xukezheng:
        #     QMessageBox.critical(self, "错误", "请填写烟草专卖许可证号码", QMessageBox.Yes)
        #     return False
        result["xukezheng"] = xukezheng
        anyou_status, anyou = self.edit_anyou.get_anyou_info()
        if not anyou_status:
            QMessageBox.critical(self, "错误", "请勾选案由", QMessageBox.Yes)
            return False
        result["anyou"] = anyou
        lingshouhu = self.edit_lingshouhu.currentText()
        result["lingshouhu"] = lingshouhu
        jvbao_time = self.edit_jvbao_time.text()
        result["jvbao_time"] = jvbao_time

        zhifarenyuan = self.edit_zhifarenyuan.text()
        if not zhifarenyuan:
            QMessageBox.critical(self, "错误", "请填写执法人员信息", QMessageBox.Yes)
            return False
        result["zhifarenyuan"] = zhifarenyuan
        xukezheng_end = self.edit_xukezheng_end.text()
        result["xukezheng_end"] = datetime.strptime(xukezheng_end, "%Y-%m-%d").date()
        location = self.edit_location.text()
        if not location:
            QMessageBox.critical(self, "错误", "请填写许可证地址", QMessageBox.Yes)
            return False
        result["location"] = location
        qita_guanxiren = self.edit_qita_guanxiren.text()
        result["qita_guanxiren"] = qita_guanxiren

        # 存储数据
        wb = openpyxl.load_workbook(self.filename)
        ws = wb["一般案件"]
        ws["A1"] = "开始时间"
        ws["B1"] = start_time
        ws["A2"] = "结束时间"
        ws["B2"] = end_time
        ws["A3"] = "被检查人(工商户字号，公司名称)"
        ws["B3"] = beijiancharen
        ws["A4"] = "联系方式"
        ws["B4"] = lianxifangshi
        ws["A5"] = "经营者姓名(法定代表人)"
        ws["B5"] = jingyingzhe
        ws["A6"] = "身份证件号码"
        ws["B6"] = shenfenzheng
        ws["A7"] = "身份证件地址"
        ws["B7"] = shenfenzheng_dizhi
        ws["A8"] = "烟草专卖许可证号码"
        ws["B8"] = xukezheng
        ws["A9"] = "案由"
        ws["B9"] = anyou[0]
        ws["C9"] = anyou[1]
        ws["D9"] = anyou[2]
        ws["E9"] = anyou[3]
        ws["F9"] = anyou[4]
        ws["A10"] = "零售户类型"
        ws["B10"] = lingshouhu
        ws["A11"] = "举报时间"
        ws["B11"] = jvbao_time
        ws["A12"] = "其他关系人"
        ws["B12"] = qita_guanxiren
        ws["A13"] = "执法人员"
        ws["B13"] = zhifarenyuan
        ws["A14"] = "许可证有效期"
        ws["B14"] = xukezheng_end
        ws["A15"] = "许可证地址"
        ws["B15"] = location
        # 存储抽样信息
        self.handle_chou_data(ws)
        wb.save(self.filename)
        createYiBan(wb, result, self.base_path)
        wb.save(self.filename)
        return True


