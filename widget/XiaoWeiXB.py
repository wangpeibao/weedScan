from datetime import datetime

import openpyxl
from PyQt5.QtWidgets import QWidget, QGridLayout, QLabel, QDateTimeEdit, QLineEdit, QTableWidget, QHeaderView, \
    QPushButton, QCheckBox, QVBoxLayout, QTableWidgetItem, QComboBox, QMessageBox

from widget.AnYou import AnYou
from widget.Function import get_yan_info, get_pinpai_and_guige, tezheng1, tezheng2, get_tezheng_info, createXiaoWei, \
    get_info_from_idcard


class XiaoWeiXB(QWidget):
    def __init__(self, base_path):
        super(XiaoWeiXB, self).__init__()
        self.filename = base_path + "案件信息.xlsx"
        self.base_path = base_path
        style_str = "QLabel{font-size: 30px;}" + "QLineEdit{font-size: 30px;}" + \
                    "QPushButton{font-size: 25px; background-color: green; min-height: 30px}" + \
                    "QComboBox{font-size: 30px;}" + "QCheckBox{font-size: 30px;}" + \
                    "QHeaderView{font-size: 25px;} QTableWidget{font-size: 25px;}" + \
                    "QDateTimeEdit{font-size: 30px;} QMessageBox{font-size: 30px;}"
        self.setStyleSheet(style_str)
        init_data = self.get_init_info()
        layout = QGridLayout()
        # 开始时间
        self.label_start_time = QLabel("开始时间")
        if init_data["start_time"]:
            self.edit_start_time = QDateTimeEdit(datetime.strptime(init_data["start_time"], "%Y-%m-%d %H:%M"))
        else:
            self.edit_start_time = QDateTimeEdit(datetime.now())
        self.edit_start_time.setDisplayFormat("yyyy-MM-dd HH:mm")
        layout.addWidget(self.label_start_time, 0, 0)
        layout.addWidget(self.edit_start_time, 0, 1)
        # 结束时间
        self.label_end_time = QLabel("结束时间")
        if init_data["end_time"]:
            self.edit_end_time = QDateTimeEdit(datetime.strptime(init_data["end_time"], "%Y-%m-%d %H:%M"))
        else:
            self.edit_end_time = QDateTimeEdit(datetime.now())
        self.edit_end_time.setDisplayFormat("yyyy-MM-dd HH:mm")
        layout.addWidget(self.label_end_time, 0, 2)
        layout.addWidget(self.edit_end_time, 0, 3)
        # 被检查人(工商户字号，公司名称)
        self.label_beijiancharen = QLabel("被检查人(工商户字号，公司名称)")
        if init_data["beijiancharen"]:
            self.edit_beijiancharen = QLineEdit(init_data["beijiancharen"])
        else:
            self.edit_beijiancharen = QLineEdit()
        layout.addWidget(self.label_beijiancharen, 1, 0)
        layout.addWidget(self.edit_beijiancharen, 1, 1)
        # 联系方式
        self.label_lianxifangshi = QLabel("联系方式")
        if init_data["lianxifangshi"]:
            self.edit_lianxifangshi = QLineEdit(init_data["lianxifangshi"])
        else:
            self.edit_lianxifangshi = QLineEdit()
        layout.addWidget(self.label_lianxifangshi, 1, 2)
        layout.addWidget(self.edit_lianxifangshi, 1, 3)
        # 经营者姓名(法定代表人)
        self.label_jingyingzhe = QLabel("经营者姓名(法定代表人)")
        if init_data["jingyingzhe"]:
            self.edit_jingyingzhe = QLineEdit(init_data["jingyingzhe"])
        else:
            self.edit_jingyingzhe = QLineEdit()
        layout.addWidget(self.label_jingyingzhe, 2, 0)
        layout.addWidget(self.edit_jingyingzhe, 2, 1)
        # 身份证件号码
        self.label_shenfenzheng = QLabel("身份证件号码")
        if init_data["shenfenzheng"]:
            self.edit_shenfenzheng = QLineEdit(init_data["shenfenzheng"])
        else:
            self.edit_shenfenzheng = QLineEdit()
        layout.addWidget(self.label_shenfenzheng, 2, 2)
        layout.addWidget(self.edit_shenfenzheng, 2, 3)
        # 身份证件地址
        self.label_shenfenzheng_dizhi = QLabel("身份证件地址")
        if init_data["shenfenzheng_dizhi"]:
            self.edit_shenfenzheng_dizhi = QLineEdit(init_data["shenfenzheng_dizhi"])
        else:
            self.edit_shenfenzheng_dizhi = QLineEdit()
        layout.addWidget(self.label_shenfenzheng_dizhi, 3, 0)
        layout.addWidget(self.edit_shenfenzheng_dizhi, 3, 1)
        # 烟草专卖许可证号码
        self.label_xukezheng = QLabel("烟草专卖许可证号码")
        if init_data["xukezheng"]:
            self.edit_xukezheng = QLineEdit(init_data["xukezheng"])
        else:
            self.edit_xukezheng = QLineEdit()
        layout.addWidget(self.label_xukezheng, 3, 2)
        layout.addWidget(self.edit_xukezheng, 3, 3)
        # 委托人姓名
        self.label_weituoren = QLabel("委托人姓名")
        if init_data["weituoren"]:
            self.edit_weituoren = QLineEdit(init_data["weituoren"])
        else:
            self.edit_weituoren = QLineEdit()
        layout.addWidget(self.label_weituoren, 4, 0)
        layout.addWidget(self.edit_weituoren, 4, 1)
        # 身份证件号码
        self.label_shenfenzheng1 = QLabel("身份证件号码")
        if init_data["shenfenzheng1"]:
            self.edit_shenfenzheng1 = QLineEdit(init_data["shenfenzheng1"])
        else:
            self.edit_shenfenzheng1 = QLineEdit()
        layout.addWidget(self.label_shenfenzheng1, 4, 2)
        layout.addWidget(self.edit_shenfenzheng1, 4, 3)
        # 案由选择
        self.label_anyou = QLabel("案由")
        layout.addWidget(self.label_anyou, 5, 0)
        self.edit_anyou = AnYou(init_data["anyou"])
        layout.addWidget(self.edit_anyou, 5, 1, 1, 3)
        self.label_location = QLabel("许可证地址")
        self.edit_location = QLineEdit()
        if init_data["location"]:
            self.edit_location.setText(init_data["location"])
        layout.addWidget(self.label_location, 6, 0)
        layout.addWidget(self.edit_location, 6, 1)
        # 分类table
        self.table_info = QTableWidget()
        self.table_info.setColumnCount(6)
        self.table_info.setHorizontalHeaderLabels(["品牌", "规格", "数量", "总价", "特征1", "特征2"])
        self.table_info.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        wb = openpyxl.load_workbook(self.filename)
        yan_datas = get_yan_info(wb)
        wb.close()
        self.setUI_table(yan_datas)
        layout.addWidget(self.table_info, 7, 0, 1, 4)
        # 完成按钮
        self.btn_finish = QPushButton("完成")
        layout.addWidget(self.btn_finish, 8, 0, 1, 4)
        # 设置格式
        self.setLayout(layout)

    # 初始化数据
    def get_init_info(self):
        wb = openpyxl.load_workbook(self.filename)
        try:
            ws = wb["小微案件"]
        except Exception as e:
            print(e)
            print("创建sheet")
            ws = wb.create_sheet(title="小微案件")
        anyou = ['', '', '', '', '']
        if ws["B11"].value:
            anyou[0] = ws["B11"].value
        if ws["C11"].value:
            anyou[1] = ws["C11"].value
        if ws["D11"].value:
            anyou[2] = ws["D11"].value
        if ws["E11"].value:
            anyou[3] = ws["E11"].value
        if ws["E12"].value:
            anyou[4] = ws["F11"].value

        data = {
            "start_time": ws["B1"].value,
            "end_time": ws["B2"].value,
            "beijiancharen": ws["B3"].value,
            "lianxifangshi": ws["B4"].value,
            "jingyingzhe": ws["B5"].value,
            "shenfenzheng": ws["B6"].value,
            "shenfenzheng_dizhi": ws["B7"].value,
            "xukezheng": ws["B8"].value,
            "weituoren": ws["B9"].value,
            "shenfenzheng1": ws["B10"].value,
            "anyou": anyou,
            "location": ws["B12"].value
        }
        wb.save(self.filename)
        return data

    # 存储输入的数据
    def handle_info(self):
        result_data = {}
        start_time = self.edit_start_time.text()
        if not start_time:
            QMessageBox.critical(self, "错误", "请填写开始时间", QMessageBox.Yes)
            return False
        result_data["start_time"] = datetime.strptime(start_time, "%Y-%m-%d %H:%M")
        end_time = self.edit_end_time.text()
        if not end_time:
            QMessageBox.critical(self, "错误", "请填写结束时间", QMessageBox.Yes)
            return False
        result_data["end_time"] = datetime.strptime(end_time, "%Y-%m-%d %H:%M")
        beijiancharen = self.edit_beijiancharen.text()
        if not beijiancharen:
            QMessageBox.critical(self, "错误", "请填写被检查人(工商户字号，公司名称)", QMessageBox.Yes)
            return False
        result_data["beijiancharen"] = beijiancharen
        lianxifangshi = self.edit_lianxifangshi.text()
        if not lianxifangshi:
            QMessageBox.critical(self, "错误", "请填写联系方式", QMessageBox.Yes)
            return False
        result_data["lianxifangshi"] = lianxifangshi
        jingyingzhe = self.edit_jingyingzhe.text()
        # if not jingyingzhe:
        #     QMessageBox.critical(self, "错误", "请填写经营者姓名(法定代表人)", QMessageBox.Yes)
        #     return False
        result_data["jingyingzhe"] = jingyingzhe
        shenfenzheng = self.edit_shenfenzheng.text()
        if shenfenzheng:
            try:
                get_info_from_idcard(shenfenzheng)
            except Exception as e:
                QMessageBox.critical(self, "错误", "请填写正确的经营者身份证号", QMessageBox.Yes)
                return False
        result_data["shenfenzheng"] = shenfenzheng
        shenfenzheng_dizhi = self.edit_shenfenzheng_dizhi.text()
        result_data["shenfenzheng_dizhi"] = shenfenzheng_dizhi
        xukezheng = self.edit_xukezheng.text()
        if not xukezheng:
            QMessageBox.critical(self, "错误", "请填写烟草专卖许可证号码", QMessageBox.Yes)
            return False
        result_data["xukezheng"] = xukezheng
        weituoren = self.edit_weituoren.text()
        result_data["weituoren"] = weituoren
        shenfenzheng1 = self.edit_shenfenzheng1.text()
        if shenfenzheng1:
            try:
                get_info_from_idcard(shenfenzheng1)
            except Exception as e:
                QMessageBox.critical(self, "错误", "请填写正确的委托人身份证号", QMessageBox.Yes)
                return False
        result_data["shenfenzheng1"] = shenfenzheng1
        location = self.edit_location.text()
        if not location:
            QMessageBox.critical(self, "错误", "请填写许可证地址", QMessageBox.Yes)
            return False
        result_data["location"] = location
        anyou_status, anyou = self.edit_anyou.get_anyou_info()
        if not anyou_status:
            QMessageBox.critical(self, "错误", "请勾选案由", QMessageBox.Yes)
            return False
        result_data["anyou"] = anyou
        # 校验都通过后－存储数据
        wb = openpyxl.load_workbook(self.filename)
        ws = wb["小微案件"]
        ws["A1"] = "开始时间"
        ws["B1"] = start_time
        ws["A2"] = "结束时间"
        ws["B2"] = end_time
        ws["A3"] = "被检查人(工商户字号，公司名称)"
        ws["B3"] = beijiancharen
        ws["A4"] = "联系方式"
        ws["B4"] = lianxifangshi
        ws["A5"] = "经营者姓名(法定代表人)"
        ws["B5"] = jingyingzhe
        ws["A6"] = "身份证件号码"
        ws["B6"] = shenfenzheng
        ws["A7"] = "身份证件地址"
        ws["B7"] = shenfenzheng_dizhi
        ws["A8"] = "烟草专卖许可证号码"
        ws["B8"] = xukezheng
        ws["A9"] = "委托人姓名"
        ws["B9"] = weituoren
        ws["A10"] = "委托人身份证件号码"
        ws["B10"] = shenfenzheng1
        ws["A11"] = "案由"
        ws["B11"] = anyou[0]
        ws["C11"] = anyou[1]
        ws["D11"] = anyou[2]
        ws["E11"] = anyou[3]
        ws["F11"] = anyou[4]
        ws["A12"] = "许可证地址"
        ws["B12"] = location
        # 处理表格数据
        self.handle_table_info(ws)
        # 最终保存
        createXiaoWei(wb, result_data, self.base_path)
        wb.save(self.filename)
        return True

    # 显示违规烟草信息的表格
    def setUI_table(self, yan_datas):
        for index, yan_data in enumerate(yan_datas):
            self.table_info.setRowCount(index + 1)
            yan_pinpai, yan_guige = get_pinpai_and_guige(yan_data["yan_name"])
            self.table_info.setItem(index, 0, QTableWidgetItem(yan_pinpai))
            self.table_info.setItem(index, 1, QTableWidgetItem(yan_guige))
            self.table_info.setItem(index, 2,
                                    QTableWidgetItem(str(yan_data["yan_count"]) + ' ' + yan_data["yan_unit"]))
            comb_box1 = QComboBox()
            comb_box1.addItems(tezheng1)
            comb_box2 = QComboBox()
            comb_box2.addItems(tezheng2)
            self.table_info.setCellWidget(index, 4, comb_box1)
            self.table_info.setCellWidget(index, 5, comb_box2)
            self.table_info.setItem(index, 3, QTableWidgetItem(yan_data["yan_total"]))

    # 存储违规烟草信息的分类
    def handle_table_info(self, ws):
        for index in range(self.table_info.rowCount()):
            ws["A" + str(index + 14)] = self.table_info.item(index, 0).text()
            ws["B" + str(index + 14)] = self.table_info.item(index, 1).text()
            ws["C" + str(index + 14)] = self.table_info.item(index, 2).text().split(' ')[0]
            ws["F" + str(index + 14)] = self.table_info.item(index, 2).text().split(' ')[1]
            ws["D" + str(index + 14)] = self.table_info.item(index, 3).text()
            ws["E" + str(index + 14)] = get_tezheng_info(
                self.table_info.cellWidget(index, 4).currentIndex(),
                self.table_info.cellWidget(index, 5).currentIndex())
