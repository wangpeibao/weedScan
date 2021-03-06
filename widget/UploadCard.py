# 上传现场照片
from PyQt5.QtCore import QTimer
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QPushButton, QGridLayout, QComboBox, QLineEdit, QMessageBox
import cv2
import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

from .UploadScene import UploadScene


class UploadCard(QWidget):
    def __init__(self, filename):
        super(UploadCard, self).__init__()
        self.filename = filename
        self.photo_index = 0  # 用于记录存储了几张图片
        self.powerStatus = False

        # 图像显示
        layout = QGridLayout()

        self.capture = None
        self.timer = QTimer(self)

        self.useBtn = QPushButton("亮证经营拍照取证")
        self.useBtn.setDisabled(True)
        self.useBtn.clicked.connect(self.useTake)

        self.noBtn = QPushButton("重新拍照")
        self.noBtn.setDisabled(True)
        self.noBtn.clicked.connect(self.restartTake)

        self.camera = QLabel("摄像头准备中。。。。")
        self.startBtn = QPushButton("打开摄像头")
        self.startBtn.clicked.connect(self.startCamera)
        self.takeBtn = QPushButton("拍照")
        self.takeBtn.clicked.connect(self.takePhoto)

        self.nextBtn = QPushButton("下一步")

        # 跳过上传许可证的备注说明
        self.remarkLable = QLabel("未亮证理由说明:")
        self.remarkComb = QComboBox()
        self.remarkComb.addItems(["无证", "其他"])
        self.edit_reason = QLineEdit()
        self.edit_reason.setPlaceholderText('选择"其他"时需要填入理由')

        layout.addWidget(self.camera, 0, 0, 1, 3)

        layout.addWidget(self.startBtn, 1, 0)
        layout.addWidget(self.takeBtn, 1, 1)
        layout.addWidget(self.noBtn, 1, 2)
        layout.addWidget(self.useBtn, 2, 0, 1, 3)
        layout.addWidget(self.remarkLable, 3, 0)
        layout.addWidget(self.remarkComb, 3, 1)
        layout.addWidget(self.edit_reason, 3, 2)
        layout.addWidget(self.nextBtn, 4, 0, 1, 3)

        self.takeBtn.setDisabled(True)
        self.noBtn.setDisabled(True)
        self.useBtn.setDisabled(True)

        self.setLayout(layout)

        # 绘制文件格式
        style_str = "QLabel{font-size: 30px;}" + "QLineEdit{font-size: 30px;}" + \
                    "QPushButton{font-size: 25px; background-color: green; min-height: 50px; min-width: 300px}" + \
                    "QComboBox{font-size: 30px;}"
        self.setStyleSheet(style_str)

    def startCamera(self):
        if not self.powerStatus:
            self.takeBtn.setDisabled(False)
            self.noBtn.setDisabled(True)
            self.useBtn.setDisabled(True)
            self.powerStatus = True
            self.startBtn.setText("关闭摄像头")
            self.capture = cv2.VideoCapture(1)
            self.timer.timeout.connect(self.display)
            self.timer.start(100)
        else:
            self.powerStatus = False
            self.startBtn.setText("打开摄像头")
            self.takeBtn.setDisabled(True)
            self.noBtn.setDisabled(True)
            self.useBtn.setDisabled(True)
            self.timer.stop()
            self.capture.release()

    # 下一步
    def release(self):
        try:
            self.timer.stop()
            self.capture.release()
        except Exception as e:
            print(e)
        # 如果没有上传过许可证，填入许可证明
        if self.photo_index == 0:
            # 判断选择无证还是其他原因
            if self.remarkComb.currentText() == "无证":
                reason = "无证"
            else:
                reason = self.edit_reason.text()
                if not reason:
                    QMessageBox.critical(self, "错误", "选择＜其他＞原因时需要填写具体原因", QMessageBox.Yes)
                    return False
            wb = openpyxl.load_workbook(self.filename)
            try:
                ws = wb["许可证照片"]
            except Exception as e:
                print(e)
                ws = wb.create_sheet(title="许可证照片")
            ws["A1"] = reason
            wb.save(self.filename)
        return True

    # 从摄像头中获取图片信息
    def get_image(self):
        ret, frame = self.capture.read()
        if ret:
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            self.pil_image = PILImage.fromarray(frame)
            height, width = frame.shape[:2]
            img = QImage(frame, width, height, QImage.Format_RGB888)
            return img
        else:
            return None

    # 显示图片
    def display(self):
        img = self.get_image()
        if img:
            img = QPixmap.fromImage(img)
            self.camera.setPixmap(img)
            self.camera.setScaledContents(True)
        else:
            pass

    # 拍照
    def takePhoto(self):
        self.timer.stop()
        self.takeBtn.setDisabled(True)
        self.useBtn.setDisabled(False)
        self.noBtn.setDisabled(False)

    # 重新拍照
    def restartTake(self):
        self.timer.start(100)
        self.takeBtn.setDisabled(False)
        self.useBtn.setDisabled(True)
        self.noBtn.setDisabled(True)

    # 使用该照片并上传
    def useTake(self):
        # 读写excel并存储图片
        self.uploadPhoto()
        self.timer.start(100)
        self.takeBtn.setDisabled(False)
        self.useBtn.setDisabled(True)
        self.noBtn.setDisabled(True)

    # 读写excel数据
    def uploadPhoto(self):  # 1现场 2许可证
        wb = openpyxl.load_workbook(self.filename)
        try:
            ws = wb["许可证照片"]
        except Exception as e:
            print(e)
            ws = wb.create_sheet(title="许可证照片")
        self.pil_image.save("./tmp.jpg")
        img_file = Image("./tmp.jpg")
        ws.add_image(img_file, "A" + str(self.photo_index * 30 + 2))
        wb.save(self.filename)
        self.photo_index += 1
