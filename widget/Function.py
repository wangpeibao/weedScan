# 一些公用的函数
import math
import re
from datetime import datetime

import openpyxl
from docxtpl import DocxTemplate

tezheng1 = [
    "A.非当地烟草专卖批发企业购进卷烟",
    "B.境外品牌卷烟(含港澳台),且未标有(由中国烟草总公司专卖)",
    "C.国产专供出口卷烟(标有\"专供出口\"字样)",
    "D.授权境外企业(含港澳台)生产卷烟(标有\"授权生产\"、\"授权出品\"等字样)",
    "E.在海关监管区内销售的未加贴秒税烟标的免税卷烟或在海关监管区外销售的加贴免税烟标的免税卷烟"
]

tezheng2 = [
    "清晰完整",
    "部分毁损",
    "全部毁损或无码"
]

anyou_list = [
    "未在当地烟草批发企业进货",
    "销售非法生产的烟草专卖品",
    "销售无标志外国卷烟",
    "销售专供出口卷烟",
    "无烟草专卖品准运证运输烟草专卖品"
]

def send_sms(code):
    from aliyunsdkcore.client import AcsClient
    from aliyunsdkcore.request import CommonRequest
    client = AcsClient('', '', 'cn-hangzhou')

    request = CommonRequest()
    request.set_accept_format('json')
    request.set_domain('dysmsapi.aliyuncs.com')
    request.set_method('POST')
    request.set_protocol_type('https')  # https | http
    request.set_version('2017-05-25')
    request.set_action_name('SendSms')

    request.add_query_param('RegionId', "cn-hangzhou")
    request.add_query_param('PhoneNumbers', "13512492836")
    request.add_query_param('SignName', "便携工具箱")
    request.add_query_param('TemplateCode', "SMS_166867869")
    request.add_query_param('TemplateParam', "{\"code\": \"%s\"}" % str(code))

    response = client.do_action_with_exception(request)
    print(str(response, encoding='utf-8'))


# 根据index获取大写字符
def getLetter(index):
    letters = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U",
               "V", "W", "X", "Y", "Z"]
    if index >= 26:
        first_letter = letters[int(index / 26) - 1]
        last_letter = letters[index % 26]
        return first_letter + last_letter
    else:
        return letters[index]


# 根据身份证号获取信息
def get_info_from_idcard(idcard):
    if idcard:
        if int(idcard[-2]) % 2 == 0:
            return "女", datetime.now().year - int(idcard[6:10])
        else:
            return "男", datetime.now().year - int(idcard[6:10])
    else:
        return "", ""


# 获取案由字符串
def get_anyou_str(ws, index):
    anyou_str = ""
    for i in range(5):
        if ws[getLetter(i + 1) + str(index)].value:
            if anyou_str:
                anyou_str += "," + ws[getLetter(i + 1) + str(index)].value
            else:
                anyou_str += ws[getLetter(i + 1) + str(index)].value
    return anyou_str


# 获取抽样记录
def get_chou_data(ws):
    index = 16
    chou_data = []
    while (True):
        yan_name = ws["B" + str(index)].value
        if not yan_name:
            break
        chou_data.append({
            "yan_pinzhong": ws["A" + str(index)].value,
            "yan_name": yan_name,
            "yan_count": ws["C" + str(index)].value,
            "yan_chou": ws["D" + str(index)].value
        })
        index += 1
    return chou_data


# html先头文件
html_headers1 = '<html><head>' +\
               '<script src="./LodopFuncs.js"></script></head><body><script>'


html_headers2 = 'function %s() { let lodop = getLodop(); lodop.PRINT_INITA(0, 0, "210mm", "297mm", "");' +\
                'lodop.SET_PRINT_MODE("PRINT_NOCOLLATE", 1);'

html_tail = 'lodop.PREVIEW();};'
html_finish = '</script></body></html>'


zhifaren_dict = {
    "郝志超": "12010112",
    "武泓伯": "12010104",
    "刘宏志": "12010118",
    "张庆": "12010116",
    "宋鹏": "12010115",
    "李根": "12010105",
    "陈亮": "12010121",
    "张昊鹏": "12010114",
    "王征": "12010120",
    "谭麒麟": "12010122",
    "张宝亮": "12010119",
    "宋治": "12010135",
    "李钢": "12010129",
    "王良辰": "12010136",
    "李涛": "12010134",
    "云亮": "12010128",
    "孟科": "12010125",
    "孟令健": "12010127",
    "沈春生": "12010117",
    "郑淼": "12010124"
}

# 生成零盒案件
def createLingHe(wb, result, base_path):
    # 数据填写部分-自愿处理记录表
    yan_data = get_yan_info(wb)
    tezheng_data = get_tezheng_data_linghe(wb)
    # 自愿处理单和勘验笔录
    for i in range(math.ceil(len(yan_data) / 10)):
        # 自愿处理
        tpl = DocxTemplate('static/零盒自愿.docx')
        end_time = result["end_time"]
        context = {
            "dangshiren": result["jingyingzhe"],
            "riqi": "%d年%d月%d日" % (end_time.year, end_time.month, end_time.day),
            "didian": result["location"],
        }
        for index in range(10):
            yan = yan_data[i*10:(i+1)*10]
            if index + 1 <= len(yan):
                context["pin%d" % index] = yan[index]["yan_pinzhong"]
                context["name%d" % index] = yan[index]["yan_name"]
                context["unit%d" % index] = yan[index]["yan_unit"]
                context["count%d" % index] = get_big_num(int(yan[index]["yan_count"]))
                context["bei%d" % index] = yan[index]["yan_id"]
            else:
                context["pin%d" % index] = ""
                context["name%d" % index] = ""
                context["unit%d" % index] = ""
                context["count%d" % index] = ""
                context["bei%d" % index] = ""
        tpl.render(context)
        tpl.save(base_path + "零盒自愿%d.docx" % i)
        # 现场勘验笔录
        tpl = DocxTemplate('static/轻微勘验.docx')
        xingbie, nianling = get_info_from_idcard(result["shenfenzheng"])
        xingbie1, nianling1 = get_info_from_idcard(result["shenfenzheng1"])
        context = {
            "syear": result["start_time"].year,
            "smonth": result["start_time"].month,
            "sday": result["start_time"].day,
            "shour": result["start_time"].hour,
            "sminute": result["start_time"].minute,
            "eyear": result["end_time"].year,
            "emonth": result["end_time"].month,
            "eday": result["end_time"].day,
            "ehour": result["end_time"].hour,
            "eminute": result["end_time"].minute,
            "location": result["location"],
            "beijiancharen": result["beijiancharen"],
            "lianxifangshi": result["lianxifangshi"],
            "jingyingzhe": result["jingyingzhe"],
            "xingbie": xingbie,
            "nianling": nianling,
            "shenfenzheng": result["shenfenzheng"],
            "shenfenzheng_dizhi": result["shenfenzheng_dizhi"],
            "xukezheng": result["xukezheng"],
            "weituoren": result["weituoren"],
            "xingbie1": xingbie1,
            "nianling1": nianling1,
            "shenfenzheng1": result["shenfenzheng1"]
        }
        self_count = 0
        self_price = 0
        for index in range(10):
            tezheng = tezheng_data[i * 10:(i + 1) * 10 - 1]
            if index + 1 <= len(tezheng):
                self_count += int(tezheng[index]["count"])
                self_price += float(tezheng[index]["price"])
                context["pin%d" % index] = tezheng[index]["pin"]
                context["gui%d" % index] = tezheng[index]["gui"]
                context["shu%d" % index] = tezheng[index]["count"]
                context["te%d" % index] = tezheng[index]["tezheng"]
                context["price%d" % index] = tezheng[index]["price"]
            else:
                context["pin%d" % index] = ""
                context["gui%d" % index] = ""
                context["shu%d" % index] = ""
                context["te%d" % index] = ""
                context["price%d" % index] = ""
        context["total_count"] = get_big_num(self_count)
        context["total_price"] = self_price
        tpl.render(context)
        tpl.save(base_path + "零盒勘验%d.docx" % i)
    # 打印封面
    tpl = DocxTemplate('static/封面.docx')
    context = {
        "beijiancharen": result["beijiancharen"],
        "location": result["location"],
        "end_date": "%d年%d月%d日" % (result["end_time"].year, result["end_time"].month, result["end_time"].day)
    }
    tpl.render(context)
    tpl.save(base_path + "封面.docx")

# 生成小微案件
def createXiaoWei(wb, result, base_path):
    html_str = html_headers1
    yan_data = get_yan_info(wb)
    tezheng_data = get_tezheng_data_xiaowei(wb)
    # 获取案由信息
    anyou_str = get_anyou_str(wb["小微案件"], 11)
    # 先保通知
    yan_top = [126, 134, 141, 149, 156, 164, 171, 179]
    yan_top += yan_top
    yan_left = [30, 67, 103, 140]
    for i in range(math.ceil(len(yan_data) / 16)):
        html_str += html_headers2 % "print_xianbaotz%d" % i
        # 当事人
        html_str += 'lodop.ADD_PRINT_TEXT("57mm","30mm", "35mm", "10mm", "%s");' % result["beijiancharen"]
        # 案由
        html_str += 'lodop.ADD_PRINT_TEXT("64mm","75mm", "75mm", "7mm", "%s");' % anyou_str
        html_str += 'lodop.ADD_PRINT_TEXT("104mm","167mm", "6mm", "7mm", "条");'
        self_num_count = 0
        self_type_count = 0
        for index, yan in enumerate(yan_data[i * 16: (i + 1) * 16]):
            self_type_count += 1
            self_num_count += int(yan["yan_count"])
            if index < 8:
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[0], yan["yan_name"])
                if len(yan["yan_name"]) > 10:
                    html_str += 'lodop.SET_PRINT_STYLEA(0, "FontSize", 5);'
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[1], yan["yan_count"] + "条")
            else:
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[2], yan["yan_name"])
                if len(yan["yan_name"]) > 10:
                    html_str += 'lodop.SET_PRINT_STYLEA(0, "FontSize", 5);'
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[3], yan["yan_count"] + "条")
        # 总类型和总数量
        num_count_big = get_big_num(self_num_count)
        type_count_big = get_big_num(self_type_count)
        html_str += 'lodop.ADD_PRINT_TEXT("187mm","93mm", "30mm", "10mm", "%s个");' % type_count_big
        html_str += 'lodop.ADD_PRINT_TEXT("187mm","158mm", "18mm", "10mm", "%s条");' % num_count_big
        html_str += html_tail
    # 勘验笔录
    for i in range(math.ceil(len(yan_data) / 10)):
        tpl = DocxTemplate('static/轻微勘验.docx')
        xingbie, nianling = get_info_from_idcard(result["shenfenzheng"])
        xingbie1, nianling1 = get_info_from_idcard(result["shenfenzheng1"])
        context = {
            "syear": result["start_time"].year,
            "smonth": result["start_time"].month,
            "sday": result["start_time"].day,
            "shour": result["start_time"].hour,
            "sminute": result["start_time"].minute,
            "eyear": result["end_time"].year,
            "emonth": result["end_time"].month,
            "eday": result["end_time"].day,
            "ehour": result["end_time"].hour,
            "eminute": result["end_time"].minute,
            "location": result["location"],
            "beijiancharen": result["beijiancharen"],
            "lianxifangshi": result["lianxifangshi"],
            "jingyingzhe": result["jingyingzhe"],
            "xingbie": xingbie,
            "nianling": nianling,
            "shenfenzheng": result["shenfenzheng"],
            "shenfenzheng_dizhi": result["shenfenzheng_dizhi"],
            "xukezheng": result["xukezheng"],
            "weituoren": result["weituoren"],
            "xingbie1": xingbie1,
            "nianling1": nianling1,
            "shenfenzheng1": result["shenfenzheng1"]
        }
        self_count = 0
        self_price = 0
        for index in range(10):
            tezheng = tezheng_data[i * 10:(i + 1) * 10 - 1]
            if index + 1 <= len(tezheng):
                self_count += int(tezheng[index]["count"])
                self_price += float(tezheng[index]["price"])
                context["pin%d" % index] = tezheng[index]["pin"]
                context["gui%d" % index] = tezheng[index]["gui"]
                context["shu%d" % index] = tezheng[index]["count"]
                context["te%d" % index] = tezheng[index]["tezheng"]
                context["price%d" % index] = tezheng[index]["price"]
            else:
                context["pin%d" % index] = ""
                context["gui%d" % index] = ""
                context["shu%d" % index] = ""
                context["te%d" % index] = ""
                context["price%d" % index] = ""
        context["total_count"] = get_big_num(self_count)
        context["total_price"] = self_price
        tpl.render(context)
        tpl.save(base_path + "小微勘验%d.docx" % i)
    # 打印封面
    tpl = DocxTemplate('static/封面.docx')
    context = {
        "beijiancharen": result["beijiancharen"],
        "location": result["location"],
        "end_date": "%d年%d月%d日" % (result["end_time"].year, result["end_time"].month, result["end_time"].day)
    }
    tpl.render(context)
    tpl.save(base_path + "封面.docx")
    # 写入到文件html中
    html_str += html_finish
    with open("C:/html_tmp/xiaowei.html", "w") as f:
        try:
            f.write(html_str)
        except Exception as e:
            print(e)
        f.close()


# 生成一般案件
def createYiBan(wb, result, base_path):
    html_str = html_headers1
    # 先保通知-----------------------------------------------------------------
    yan_data = get_yan_info(wb)
    has_card = wb["许可证照片"]["A1"].value
    ws = wb["一般案件"]
    # 获取案由信息
    anyou_str = get_anyou_str(ws, 9)
    yan_top = [120, 128, 135, 142, 150, 158, 165, 173]
    yan_top += yan_top
    yan_left = [32, 69, 105, 141]
    for i in range(math.ceil(len(yan_data) / 16)):
        html_str += html_headers2 % "print_xianbaotz%d" % i
        # 当事人
        html_str += 'lodop.ADD_PRINT_TEXT("52mm","30mm", "35mm", "10mm", "%s");' % result["beijiancharen"]
        # 案由
        html_str += 'lodop.ADD_PRINT_TEXT("61mm","73mm", "73mm", "10mm", "%s");' % anyou_str
        html_str += 'lodop.ADD_PRINT_TEXT("99mm","167mm", "6mm", "7mm", "条");'
        self_num_count = 0
        self_type_count = 0
        for index, yan in enumerate(yan_data[i * 16: (i + 1) * 16]):
            self_type_count += 1
            self_num_count += int(yan["yan_count"])
            if index < 8:
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[0], yan["yan_name"])
                if len(yan["yan_name"]) > 10:
                    html_str += 'lodop.SET_PRINT_STYLEA(0, "FontSize", 5);'
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[1], yan["yan_count"] + "条")
            else:
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[2], yan["yan_name"])
                if len(yan["yan_name"]) > 10:
                    html_str += 'lodop.SET_PRINT_STYLEA(0, "FontSize", 5);'
                html_str += 'lodop.ADD_PRINT_TEXT("%dmm","%dmm", "35mm", "7mm", "%s");' % \
                            (yan_top[index], yan_left[3], yan["yan_count"] + "条")
        # 总类型和总数量
        num_count_big = get_big_num(self_num_count)
        type_count_big = get_big_num(self_type_count)
        html_str += 'lodop.ADD_PRINT_TEXT("182mm","79mm", "25mm", "10mm", "%s");' % (type_count_big + "个")
        html_str += 'lodop.ADD_PRINT_TEXT("182mm","140mm", "30mm", "10mm", "%s");' % (num_count_big + "条")
        html_str += html_tail
    # 勘验笔录 ----------------------------------------------------------------
    tpl = DocxTemplate('static/一般勘验.docx')
    start_time = result["start_time"]
    xingbie, nianling = get_info_from_idcard(result["shenfenzheng"])
    zhifaren1 = result["zhifarenyuan"].split("-")[0]
    if zhifaren1 in zhifaren_dict.keys():
        zhifaren1 += "(" + zhifaren_dict[zhifaren1] + ")"
    else:
        zhifaren1 += "(           )"
    zhifaren2 = result["zhifarenyuan"].split("-")[1]
    if zhifaren2 in zhifaren_dict.keys():
        zhifaren2 += "(" + zhifaren_dict[zhifaren2] + ")"
    else:
        zhifaren2 += "(           )"
    zhifaren_count = result["zhifarenyuan"].split("-")[2]
    tiao_count = 0
    fei_count = 0
    posun_count = 0
    posun_str = ""
    peisong_count1 = 0
    peisong_count2 = 0
    peisong_count3 = 0
    peisong_str1 = ""
    peisong_code = ""
    peisong_str2 = ""
    peisong_str3 = ""
    jia_str = ""
    jia_count = 0
    si_str1 = ""
    si_str_code = ""
    si_str2 = ""
    si_str3 = ""
    si_count1 = 0
    si_count2 = 0
    si_count3 = 0
    for yan in yan_data:
        tiao_count += int(yan["yan_count"])
        # 破损的判断
        if yan["yan_baozhuang"] == "有破损":
            posun_count += int(yan["yan_count"])
            if posun_str:
                posun_str = "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
            else:
                posun_str = yan["yan_name"] + str(yan["yan_count"]) + "条"
        # 配送码状态
        if yan["yan_sort"] == "非":
            fei_count += int(yan["yan_count"])
            if yan["yan_peisong_status"] == "有配送码":
                peisong_count1 += int(yan["yan_count"])
                if peisong_str1:
                    peisong_str1 += "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
                else:
                    peisong_str1 += yan["yan_name"] + str(yan["yan_count"]) + "条"
                if peisong_code:
                    if yan["yan_peisong_code"] not in peisong_code:
                        peisong_code += "、" + yan["yan_peisong_code"]
                else:
                    peisong_code = yan["yan_peisong_code"]
            if yan["yan_peisong_status"] == "无配送码":
                peisong_count2 += int(yan["yan_count"])
                if peisong_str2:
                    peisong_str2 += "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
                else:
                    peisong_str2 += yan["yan_name"] + str(yan["yan_count"]) + "条"
            if yan["yan_peisong_status"] == "配送码模糊不清":
                peisong_count3 += int(yan["yan_count"])
                if peisong_str3:
                    peisong_str3 += "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
                else:
                    peisong_str3 += yan["yan_name"] + str(yan["yan_count"]) + "条"
        # 假的状态
        if yan["yan_sort"] == "假":
            jia_count += int(yan["yan_count"])
            if jia_str:
                jia_str += "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
            else:
                jia_str += yan["yan_name"] + str(yan["yan_count"]) + "条"
        # 私的状态
        if yan["yan_sort"] == "无专卖字样":
            si_count1 += int(yan["yan_count"])
            if si_str1:
                si_str1 = "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
            else:
                si_str1 = yan["yan_name"] + str(yan["yan_count"]) + "条"
            if si_str_code:
                if yan["yan_id"] not in si_str_code:
                    si_str_code += "、" + yan["yan_id"][:3]
            else:
                si_str_code = yan["yan_id"][:3]
        if yan["yan_sort"] == "授权生产":
            si_count2 += int(yan["yan_count"])
            if si_str2:
                si_str2 = "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
            else:
                si_str2 = yan["yan_name"] + str(yan["yan_count"]) + "条"
        if yan["yan_sort"] == "专供出口":
            si_count3 += int(yan["yan_count"])
            if si_str3:
                si_str3 = "、" + yan["yan_name"] + str(yan["yan_count"]) + "条"
            else:
                si_str3 = yan["yan_name"] + str(yan["yan_count"]) + "条"
    kanyan_str = ""
    if result["jvbao_time"] == "2000-01-01 00:00":
        kanyan_str += "%d年%d月%d日%d时%d分，" % (start_time.year, start_time.month, start_time.day, start_time.hour, start_time.minute)
        kanyan_str += "天津市区第一烟草专卖局%s、%s等%s名行政执法人员，" % (zhifaren1, zhifaren2, zhifaren_count)
    else:
        jvbao_time = datetime.strptime(result["jvbao_time"], "%Y-%m-%d %H:%M")
        kanyan_str += "%d年%d月%d日%d时%d分，" % (jvbao_time.year, jvbao_time.month, jvbao_time.day, jvbao_time.hour, jvbao_time.minute)
        kanyan_str += "天津市区第一烟草专卖局接到群众举报后,随即派遣%s、%s等%s名行政执法人员" % (zhifaren1, zhifaren2, zhifaren_count)
        kanyan_str += "于%d年%d月%d日%d时%d分，" % (start_time.year, start_time.month, start_time.day, start_time.hour, start_time.minute)
    kanyan_str += "来到%s(%s)，在出示了执法证件亮明身份后，" % (result["location"], result["beijiancharen"])
    kanyan_str += "对其经营场所进行了检查，当场查获当事人%s在其经营场所内摆卖的涉嫌违法卷烟" % result["beijiancharen"]
    kanyan_str += "共计%d个品种%d条，" % (len(yan_data), tiao_count)
    # 处理破损数据
    if posun_count == tiao_count:
        kanyan_str += "包装全部破损。"
    elif posun_count == 0:
        kanyan_str += "包装全部完好无破损。"
    else:
        kanyan_str += "其中，%s，共%d条卷烟有破损，其余卷烟包装完好无破损。" % (posun_str, posun_count)
    # 是否是个体
    if result["lingshouhu"] == "个体零售户":
        if result["qita_guanxiren"] == "" or result["qita_guanxiren"] == "店员":
            kanyan_str += "经与当事人%s一起现场勘验，其中：" % result["jingyingzhe"]
        else:
            kanyan_str += "经与当事人%s的%s一起现场勘验，其中：" % (result["jingyingzhe"], result["qita_guanxiren"])
    else:
        if result["qita_guanxiren"] == "" or result["qita_guanxiren"] == "店员":
            kanyan_str += "经与当事人%s的法定代表人%s一起现场勘验，其中：" % (result["beijiancharen"], result["jingyingzhe"])
        else:
            kanyan_str += "经与当事人%s的%s一起现场勘验，其中：" % (result["beijiancharen"], result["qita_guanxiren"])
    # 处理配送码
    if peisong_count1 > 0:  # 全部打有编码
        kanyan_str += '%s，共%d条卷烟的条包上分别打有第二行以"%s"开头的32位配送编码，与当事人零售客户代码不符。' % (peisong_str1, peisong_count1, peisong_code)
    if peisong_count2 > 0:  # 全部未打有
        kanyan_str += '%s，共%d条卷烟的条包上均未打有32位配送编码。' % (peisong_str2, peisong_count2)
    if peisong_count3 > 0:  # 全部模糊
        kanyan_str += '%s，共%d条卷烟的条包上打有的32位配送编码均模糊无法识别。'
    if fei_count:
        kanyan_str += "上述%d条卷烟初步认定为未在当地烟草批发企业进货的卷烟。" % fei_count
    # 假
    kanyan_str += "%s，共%d条卷烟的外包装特征与真品卷烟的外包装特征不符，初步认定为非法生产的卷烟。" % (jia_str, jia_count)
    # 私
    if si_count1 > 0:
        kanyan_str += "%s，共%d条卷烟的包装上分别打有“%s”开头的条形码且均未打有32位配送编码，并且其包装上均未标注有“由中国烟草总公司专卖”字样。" % (si_str1, si_count1, si_str_code)
    if si_count2 > 0:
        kanyan_str += "%s, 共%d条卷烟的包装上均标注有“授权生产”字样。" % (si_str2, si_count2)
    if si_count1 + si_count2 > 0:
        kanyan_str += "上述%d条卷烟初步认定为无标志外国卷烟。" % (si_count1 + si_count2)
    # 专供出口
    if si_count3 > 0:
        kanyan_str += "%s, 共%d条卷烟的包装上均打有以“6901028”开头的条形码且均未打有32位配送编码，并且其包装上均标准有“专供出口”字样，初步认定为专供出口卷烟。" % (si_str3, si_count3)
    # 结尾部分
    if result["lingshouhu"] == "个体零售户":
        if result["qita_guanxiren"] == "" or result["qita_guanxiren"] == "店员":
            kanyan_str += "当事人%s在案发现场，" % result["jingyingzhe"]
        else:
            kanyan_str += "当事人%s不在案发现场，" % result["jingyingzhe"]
            kanyan_str += "其%s当时在案发现场，" % result["qita_guanxiren"]
    else:
        if result["qita_guanxiren"] == "" or result["qita_guanxiren"] == "店员":
            kanyan_str += "当事人%s的法定代表人%s在案发现场，" % (result["beijiancharen"], result["jingyingzhe"])
        else:
            kanyan_str += "当事人%s的法定代表人%s不在案发现场，" % (result["beijiancharen"], result["jingyingzhe"])
            kanyan_str += "其%s当时在案发现场，" % result["qita_guanxiren"]
    if result['xukezheng']:
        kanyan_str += "当事人%s持有烟草专卖零售许可证，" % result["beijiancharen"]
        kanyan_str += "证号是%s，有效期至%d年%d月%d日。" % (result['xukezheng'], result["xukezheng_end"].year, result["xukezheng_end"].month, result["xukezheng_end"].day)
    else:
        kanyan_str += "当事人%s不持有烟草专卖零售许可证。" % result["beijiancharen"]
    kanyan_str += "当事人无法提供上述卷烟的天津市烟草公司第一分公司开具的有效进货证明，我局行政执法人员依法对上述卷烟予以先行登记保存（先行登记保存通知书编号为一                ）。"

    context = {
        "syear": result["start_time"].year,
        "smonth": result["start_time"].month,
        "sday": result["start_time"].day,
        "shour": result["start_time"].hour,
        "sminute": result["start_time"].minute,
        "eday": result["end_time"].day,
        "ehour": result["end_time"].hour,
        "eminute": result["end_time"].minute,
        "location": result["location"],
        "beijiancharen": result["beijiancharen"],
        "lianxifangshi": result["lianxifangshi"],
        "jingyingzhe": result["jingyingzhe"],
        "xingbie": xingbie,
        "nianling": nianling,
        "shenfenzheng": result["shenfenzheng"],
        "shenfenzheng_dizhi": result["shenfenzheng_dizhi"],
        "xukezheng": result["xukezheng"],
        "kanyan_str": kanyan_str
    }
    tpl.render(context)
    tpl.save(base_path + "一般勘验.docx")
    # 处理抽样记录表
    chou_datas = get_chou_data(wb["一般案件"])
    wb = openpyxl.load_workbook('static/一般抽样.xlsx')
    ws = wb["抽样"]
    ws["B3"] = result["beijiancharen"]
    ws["B4"] = "当事人涉嫌" + anyou_str
    ws["B5"] = "%d年%d月%d日%d时" % (result["end_time"].year, result["end_time"].month, result["end_time"].day, result["end_time"].hour)
    ws["B6"] = result["location"]
    for i in range(math.ceil(len(chou_datas) / 15)):
        chou_data = chou_datas[i * 15 : (i + 1) * 15]
        for index, chou in enumerate(chou_data):
            ws["A" + str(index + 9)] = chou["yan_pinzhong"]
            ws["B" + str(index + 9)] = chou["yan_name"]
            ws["C" + str(index + 9)] = chou["yan_count"]
            ws["D" + str(index + 9)] = chou["yan_chou"]
        wb.save(base_path + "一般抽样%d.xlsx" % i)

    # 打印封面
    tpl = DocxTemplate('static/封面.docx')
    context = {
        "beijiancharen": result["beijiancharen"],
        "location": result["location"],
        "end_date": "%d年%d月%d日" % (result["end_time"].year, result["end_time"].month, result["end_time"].day)
    }
    tpl.render(context)
    tpl.save(base_path + "封面.docx")
    # 写入到文件html中
    html_str += html_finish
    with open("C:/html_tmp/yiban.html", "w") as f:
        try:
            f.write(html_str)
        except Exception as e:
            print(e)
        f.close()


# 零盒案件的特征信息
def get_tezheng_data_linghe(wb):
    ws = wb["零盒案件"]
    start_index = 14
    res = []
    while(ws["A" + str(start_index)].value):
        res.append({
            "pin": ws["A" + str(start_index)].value,
            "gui": ws["B" + str(start_index)].value,
            "count": ws["C" + str(start_index)].value,
            "price": ws["D" + str(start_index)].value,
            "tezheng": ws["E" + str(start_index)].value
        })
        start_index += 1
    return res

# 小微案件的特征信息
def get_tezheng_data_xiaowei(wb):
    ws = wb["小微案件"]
    start_index = 14
    res = []
    while(ws["A" + str(start_index)].value):
        res.append({
            "pin": ws["A" + str(start_index)].value,
            "gui": ws["B" + str(start_index)].value,
            "count": ws["C" + str(start_index)].value,
            "price": ws["D" + str(start_index)].value,
            "tezheng": ws["E" + str(start_index)].value
        })
        start_index += 1
    return res





# 根据文件名获取烟草信息
def get_yan_info(wb):
    try:
        ws = wb["违规烟草记录"]
    except Exception as e:
        print(e)
        return []
    results = []
    for row in ws.rows:
        if row[0].value == "条形码":
            results = []
        else:
            results.append({
                "yan_id": row[0].value,
                "yan_name": row[1].value,
                "yan_price": row[2].value,
                "yan_unit": row[3].value,
                "yan_sort": row[4].value,
                "yan_count": row[5].value,
                "yan_total": row[6].value,
                "yan_pinzhong": row[7].value,
                "yan_baozhuang": row[8].value,
                "yan_peisong_status": row[9].value,
                "yan_peisong_code": row[10].value
            })
    return results


# 获取抽样烟草信息
def get_chouyang_yaninfo(wb):
    ws = wb["一般案件-抽样清单"]
    result = []
    index = 6
    while(True):
        if ws["A" + str(index)]:
            result.append({
                "wupin_name": ws["A" + str(index)].value,
                "yan_name": ws["B" + str(index)].value,
                "yan_count": ws["C" + str(index)].value,
                "chou_count": ws["D" + str(index)].value
            })
        else:
            break
    return result


# 根据烟草的品规获取品牌和规格
def get_pinpai_and_guige(yan_name):
    if "（" in yan_name:
        yan_split = yan_name.split("（")
        yan_pinpai = yan_split[0]
        if len(yan_split) > 1:
            yan_guige = yan_split[1].split("）")[0]
        else:
            yan_guige = ""
    else:
        yan_split = yan_name.split("(")
        yan_pinpai = yan_split[0]
        if len(yan_split) > 1:
            yan_guige = yan_split[1].split(")")[0]
        else:
            yan_guige = ""
    return yan_pinpai, yan_guige


# 获取特征信息
def get_tezheng_info(tezheng1, tezheng2):
    tezheng_list = ["A", "B", "C", "D", "E", "F"]
    return tezheng_list[tezheng1] + str(tezheng2 + 1)


# 获取数字的大写字符串
def get_big_num(num):
    big_num = ["零", "壹", "贰", "叁", "肆", "伍", "陆", "柒", "捌", "玖"]
    big_unit = ["拾", "佰", "仟", "万"]
    num_str = ""
    if num / 10000 > 1:
        num_str += big_num[int(num / 10000)] + big_unit[3]
        num = num % 10000
    if num / 1000 > 1:
        num_str += big_num[int(num / 1000)] + big_unit[2]
        num = num % 1000
    if num / 100 > 1:
        if num_str and num_str[-1] != big_unit[2]:
            num_str += big_num[0]
        num_str += big_num[int(num / 100)] + big_unit[1]
        num = num % 100
    if num / 10 > 1:
        if num_str and num_str[-1] != big_unit[1]:
            num_str += big_num[0]
        num_str += big_num[int(num / 10)] + big_unit[0]
        num = num % 10
    if num > 0:
        if num_str and num_str[-1] != big_unit[0]:
            num_str += big_num[0]
        num_str += big_num[num]
    return num_str


# 获取一般案件的勘验笔录
def get_kanyan_str(wb):
    location = wb["基础信息"]["B2"].value
    if wb["许可证照片"]["A1"].value == "无证":
        has_card = False
    else:
        has_card = True
    ws = wb["一般案件"]
    results = get_yan_info(wb)
    count_data = {
        "jia": {"count": 0, "price": 0, "data": []},
        "si": {"count": 0, "price": 0, "data": []},
        "fei": {"count": 0, "price": 0, "data": []},
        "count": 0,
        "price": 0,
        "type": 0,
        "is_tiao": False,
    }
    for index, result in enumerate(results):
        # 统计数据
        count_data["type"] += 1
        count_data["count"] += int(result["yan_count"])
        count_data["price"] += float(result["yan_total"])
        if result["yan_sort"] == "假":
            count_data["jia"]["count"] += int(result["yan_count"])
            count_data["jia"]["price"] += float(result["yan_total"])
            count_data["jia"]["data"].append(result)
        elif result["yan_sort"] == "私":
            count_data["si"]["count"] += int(result["yan_count"])
            count_data["si"]["price"] += float(result["yan_total"])
            count_data["si"]["data"].append(result)
        else:
            count_data["fei"]["count"] += int(result["yan_count"])
            count_data["fei"]["price"] += float(result["yan_total"])
            count_data["fei"]["data"].append(result)
        if "条" in result["yan_unit"]:
            count_data["is_tiao"] = True
    kanyan = ""
    if ws["B11"].value == "2000-01-01 00:00":  # 非举报
        kanyan += ws["B12"].value + "，天津市区第一烟草专卖局行政执法人员" + ws["B13"].value + "进行市场检查时，"
    else:
        kanyan += ws["B11"].value + "，天津市区第一烟草专卖局接到群众举报后，随即派遣" + ws["B13"].value + "行政执法人员于"
        kanyan += ws["B12"].value + "，"
    kanyan += location + "，在出示了执法证件亮明身份后，对其经营场所进行了检查,当场查获当事人" + ws["B3"].value
    kanyan += "在其经营场所内摆卖的涉嫌违法卷烟共计%d个品种卷烟%d条，" % (count_data["type"], count_data["count"])
    if ws["B10"].value == "个体零售户":
        kanyan += "包装全部完好无破损。经与当事人" + ws["B3"].value + "一起现场勘验，其中："
    else:
        kanyan += "包装全部完好无破损。经与当事人%s的%s一起现场勘验，其中：" % (ws["B3"].value, ws["B5"].value)
    # 违规烟草信息
    if count_data["fei"]["data"]:
        for self_data in count_data["fei"]["data"]:
            kanyan += self_data["yan_name"] + "、" + str(self_data["yan_count"]) + self_data["yan_unit"] + "，"
        kanyan += "共%s条卷烟条包打有的32位配送编码均模糊不清无法识别，" % str(count_data["fei"]["count"])
        kanyan += "上述%s条卷烟初步认定为未在当地烟草批发企业进货的卷烟；" % str(count_data["fei"]["count"])
    if count_data["si"]["data"]:
        for self_data in count_data["si"]["data"]:
            kanyan += self_data["yan_name"] + "、" + str(self_data["yan_count"]) + self_data["yan_unit"] + "，"
        kanyan += "共%s条卷烟条包上均未打有32位配送编码，" % str(count_data["si"]["count"])
        kanyan += "并且其条包上均未标注有“由中国烟草总公司专卖”字样，初步认定为无标志外国卷烟；"
    if count_data["jia"]["data"]:
        for self_data in count_data["jia"]["data"]:
            kanyan += self_data["yan_name"] + "、" + str(self_data["yan_count"]) + self_data["yan_unit"] + "，"
        kanyan += "共%s条卷烟条包特征与真品卷烟不符，初步认定为假冒卷烟。" % str(count_data["jia"]["count"])
    if ws["B10"].value == "个体零售户":
        if has_card:  # 是否有证
            kanyan += "当事人%s当时在案发现场，并持有烟草专卖零售许可证，" % ws["B3"].value
            kanyan += "证号是%s，" % ws["B8"].value
            kanyan += "有效期至%s。" % ws["B14"].value
            kanyan += "当事人%s无法提供上述卷烟的天津市烟草公司第一分公司开具的有效进货证明，" % ws["B3"].value
            kanyan += "我局行政执法人员依法对上述卷烟予以先行登记保存（先行登记保存通知书编号为%s）。" % ws["B15"].value
        else:  # 无证
            kanyan += "当事人%s当时在案发现场。" % ws["B3"].value
            kanyan += "当事人%s无烟草专零售许可证，无法提供该批卷烟的准运证及其他有效的购货凭证，" % ws["B3"].value
            kanyan += "我局行政执法人员依法对上述卷烟予以先行登记保存（先行登记保存通知书编号为%s）。" % ws["B15"].value
    else:
        kanyan += "当事人%s持有烟草专卖零售许可证，" % ws["B3"].value
        kanyan += "证号是%s，" % ws["B8"].value
        kanyan += "有效期至%s。" % ws["B14"].value
        kanyan += "当事人%s的%s当时在案发现场，" % (ws["B3"].value, ws["B5"].value)
        kanyan += "但其无法提供上述卷烟的天津市烟草公司第一分公司开具的有效进货证明，"
        kanyan += "我局行政执法人员依法对上述卷烟予以先行登记保存（先行登记保存通知书编号为%s）。" % ws["B15"].value
    return kanyan



if __name__ == "__main__":
    send_sms(1234)

